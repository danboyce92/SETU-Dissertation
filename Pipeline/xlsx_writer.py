import openpyxl

def write_run_data(xlsx_path: str, sheet_name: str, run_number: int, values: dict) -> None:

    wb = openpyxl.load_workbook(xlsx_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Map header text (row 1) -> column index
    header_col = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            header_col[str(header).strip()] = col

    # Find the row in column A
    target_label = f"Run {run_number}"
    row_idx = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value).strip() == target_label:
            row_idx = row
            break
    if row_idx is None:
        raise ValueError(f"Could not find row '{target_label}' in sheet '{sheet_name}'.")

    for header, value in values.items():
        col = header_col.get(header)
        if col is None:
            print(f"WARNING: column '{header}' not found in sheet '{sheet_name}', skipping.")
            continue
        ws.cell(row=row_idx, column=col, value=value)

    wb.save(xlsx_path)
    print(f"Wrote {len(values)} value(s) to '{sheet_name}' -> {target_label}")
